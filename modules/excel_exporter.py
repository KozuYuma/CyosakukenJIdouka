"""
Excel 出力モジュール
openpyxl で整形済み Excel ファイルを生成する
"""
import io
import re

import pandas as pd
from openpyxl import Workbook
from modules.matcher import duration_to_min_sec
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# --- 申告フォーマット用コード整形ヘルパー ---

def _fmt_jasrac_code(val) -> str:
    """JASRAC作品コードを XXX-XXXX-X(X) 形式に整形する。"""
    if not val or (isinstance(val, float) and pd.isna(val)):
        return ""
    cleaned = re.sub(r"[-\s]", "", str(val)).upper().strip()
    if len(cleaned) < 7:
        return cleaned
    return cleaned[:3] + "-" + cleaned[3:7] + "-" + cleaned[7:]


def _fmt_cd_number(val) -> str:
    """CD番号を 英字3〜4文字-数字 形式に整形する。スペース・ハイフン・なし を統一してハイフンに。"""
    if not val or (isinstance(val, float) and pd.isna(val)):
        return ""
    s = str(val).strip()
    if not s:
        return ""
    # 英字3〜4文字 + 任意の空白/ハイフン + 数字以降 → 統一フォーマット
    m = re.match(r"^([A-Za-z]{3,4})[\s\-]*(\d.*)$", s)
    if m:
        return m.group(1).upper() + "-" + m.group(2)
    return s


# --- 確認ステータスの色定義（背景色 RGB 16進） ---
STATUS_COLOR_MAP: dict[str, str] = {
    "未調査":           "FFFFFF",  # 白
    "作曲者一致":       "DCEDC8",  # 薄黄緑（曲名＋作曲者が一致。候補ありより確度が高い）
    "アーティスト一致": "F0F4C3",  # 薄黄緑（淡）（曲名＋アーティストが一致。作曲者一致より弱い）
    "台帳一致":         "E3F2FD",  # 薄青（管理番号で自社の台帳に当たった）
    "台帳一致（曲名）": "EDF4FB",  # 薄青（淡）（トラック番号＋曲名で当たった）
    "候補あり":         "FFF9C4",  # 薄黄
    "複数候補あり":     "FFE082",  # 濃い黄（要確認）
    "確定":             "C8E6C9",  # 薄緑
    "要確認":           "FFCDD2",  # 薄赤
    "J-WID要確認":      "FFE0B2",  # 薄オレンジ
    "NexTone要確認":    "F3E5F5",  # 薄紫
    "ライブラリ元確認": "E1F5FE",  # 薄水色
    "Audiostock確認":  "F1F8E9",  # 薄黄緑
    "MP3補助確認":      "FFF3E0",  # 薄アンバー
}

HEADER_FILL = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
DEFAULT_FONT = Font(size=10)

# Excel 文字列強制列（数値として解釈させない列名）
STRING_FORCE_COLS = {
    "JASRAC作品コード",
    "NexTone管理番号",
    "CD番号",
    "元管理番号",
    "ライブラリ盤番号",
    "トラック番号",
    "Audiostock管理番号",
}


def _write_sheet(ws, df: pd.DataFrame) -> None:
    """DataFrame を 1 ワークシートに書き込む（ヘッダー固定・フィルター・列幅調整つき）"""
    if df is None or len(df) == 0:
        ws.cell(row=1, column=1, value="データなし")
        return

    col_names = list(df.columns)

    # ---- ヘッダー行（1行目） ----
    for col_idx, col_name in enumerate(col_names, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ---- データ行 ----
    status_col_idx = col_names.index("確認ステータス") + 1 if "確認ステータス" in col_names else None

    for row_idx, (_, series) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(col_names, 1):
            val = series[col_name]

            # NaN・None を空文字に
            if pd.isna(val):
                val = ""

            # 文字列強制列はアポストロフィを付けて文字列として保存
            if col_name in STRING_FORCE_COLS:
                val = str(val) if val != "" else ""

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DEFAULT_FONT

        # 確認ステータス列の背景色
        if status_col_idx:
            status_val = str(series.get("確認ステータス", ""))
            color = STATUS_COLOR_MAP.get(status_val, "FFFFFF")
            ws.cell(row=row_idx, column=status_col_idx).fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )

    # ---- フィルター & ウィンドウ固定 ----
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # ---- 列幅自動調整（最大 50 文字） ----
    for col_idx, col_name in enumerate(col_names, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col_name))
        sample_rows = min(ws.max_row, 200)  # サンプル 200 行で計算
        for r in range(2, sample_rows + 1):
            val = ws.cell(row=r, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 50))
        ws.column_dimensions[col_letter].width = max_len + 2


def _write_notes_sheet(ws) -> None:
    """処理ルール・凡例を確認メモシートに書き込む"""
    notes = [
        ("【処理ルール】", ""),
        ("", "・NUENDOの使用尺は編集後の使用尺です。曲のフル尺ではありません。"),
        ("", "・曲特定にはプロジェクト Audio フォルダ内の WAV フル尺を使います。"),
        ("", "・WAV で特定できない場合のみ MP3 を補助として使います。"),
        ("", "・MP3 タグ情報は前提にしません（タグなしファイルにも対応）。"),
        ("", "・NUENDO のファイル名より、イベント名を原曲特定の優先キーとします。"),
        ("", ""),
        ("【照合優先順位】", ""),
        ("", "1. イベント名と WAV ファイル名の完全一致"),
        ("", "2. イベント名から拡張子・余分な記号を除いた正規化一致"),
        ("", "3. 管理番号一致"),
        ("", "4. 曲名一致（管理番号除去後）"),
        ("", "5. NUENDO ファイル名一致"),
        ("", "6. 部分一致"),
        ("", ""),
        ("【管理番号形式】", ""),
        ("", "ライブラリ管理番号: 数字1桁 + 英字2文字 - 3桁 - 2桁"),
        ("", "  例: 6ST-653-09 GO! GO!  →  管理番号: 6ST-653-09 / 盤番号: 6ST-653 / トラック: 09"),
        ("", "  有効系列: 1ST〜7ST / 1AN〜5AN / 1VO〜2VO / 1VJ〜2VJ"),
        ("", ""),
        ("", "Audiostock 管理番号: audiostock_数字"),
        ("", "  例: audiostock_856447_残念なシーンのジングル  →  番号: 856447"),
        ("", ""),
        ("【確認ステータス凡例】", ""),
        ("未調査",           "まだ調査していない"),
        ("作曲者一致",       "曲名に加えて作曲者名も一致（候補ありより確度が高い）"),
        ("アーティスト一致", "曲名に加えてアーティスト名が一致（カバー音源の可能性が残る）"),
        ("台帳一致",         "管理番号で自社CD台帳・共有楽曲データに一致"),
        ("台帳一致（曲名）", "管理番号が無く、トラック番号＋曲名で自社CD台帳に一致"),
        ("候補あり",         "候補は特定できているが確認中"),
        ("複数候補あり",     "候補が複数あり、どれか選ぶ必要がある"),
        ("確定",             "権利情報確定"),
        ("要確認",           "何らかの問題あり"),
        ("J-WID要確認",      "J-WIDでの確認が必要"),
        ("NexTone要確認",    "NexToneでの確認が必要"),
        ("ライブラリ元確認", "ライブラリ元への確認が必要"),
        ("Audiostock確認",  "Audiostockでの確認が必要"),
        ("MP3補助確認",      "WAVで照合できず、MP3での補助確認が必要"),
    ]

    bold_font = Font(bold=True, size=10)
    normal_font = Font(size=10)

    for row_idx, (col_a, col_b) in enumerate(notes, 1):
        cell_a = ws.cell(row=row_idx, column=1, value=col_a)
        ws.cell(row=row_idx, column=2, value=col_b)
        if col_a.startswith("【") or col_a in STATUS_COLOR_MAP:
            cell_a.font = bold_font
            if col_a in STATUS_COLOR_MAP:
                color = STATUS_COLOR_MAP[col_a]
                cell_a.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        else:
            cell_a.font = normal_font

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 65


# 申告フォーマット: songs_df 列名 → 申告列名 マッピング
_SHINKOK_RENAME: dict[str, str] = {
    "曲名":             "楽曲名",
    "使用形態":         "使用形態",
    "音源区分":         "音源区分",
    "レコード会社名":   "レコード会社名",
    "CD番号":           "レコード番号",
    "I/V区分":         "I/V区分",
    "邦洋区分":         "邦・洋区分",
    "原訳詞区分":       "原・訳詞区分",
    "JASRAC作品コード": "JASRACコード",
    "NexTone管理番号":  "NexTone管理番号",
    "作詞者":           "作詞",
    "作曲者":           "作曲",
    "編曲者":           "編曲",
    "アーティスト":     "アーティスト",
    "訳詞者":           "訳詞",
    "自社楽曲ID":       "自社楽曲ID",
}

# 申告列の並び順（参照列は末尾）
_SHINKOK_ORDER = [
    "使用形態", "楽曲名", "音源区分", "レコード会社名", "レコード番号",
    "使用時間（分）", "使用時間（秒）",
    "I/V区分", "邦・洋区分", "原・訳詞区分",
    "JASRACコード", "NexTone管理番号", "作詞", "作曲", "アーティスト", "編曲", "訳詞",
    "自社楽曲ID",
    # 参照列（申告書には不要だが確認用）。
    # トラック（M1-1 など）は NUENDO の作りの都合で、申告には使わないので出さない
    "イベント名", "元管理番号", "START TIME", "使用尺",
    # 補助列（申告書外・右端）。放送・配信は J-WID の管理状況の同名項目で、
    # 委任者と並べて権利まわりをひとまとまりに見せる
    "確認ステータス", "委任者", "放送", "配信", "CD名", "正式タイトル",
]

#: 申告書には出ないが、確認のために持ち回る楽曲側の欄。
#: 元管理番号は自社CDの台帳を引くための鍵。Cue に番号が書かれて
#: いない曲は空になるので、表で手で足せるように出しておく
#: 正式タイトルは MusicBrainz / Spotify が尺つきで当てた曲名。
#: 一括検索がどの名前で引いたのかを表で見えるようにするために出す
_SHINKOK_AUX_COLS = ("確認ステータス", "委任者", "放送", "配信", "CD名",
                     "元管理番号", "正式タイトル")

# 申告シートで整数として出力する列
_SHINKOK_INT_COLS = {"使用時間（分）", "使用時間（秒）"}

# 参照列（申告シート内でグレー背景にする）
_SHINKOK_REF_COLS = {"イベント名", "START TIME", "使用尺", *_SHINKOK_AUX_COLS}


def build_shinkok_df(songs_df: pd.DataFrame, events_df: pd.DataFrame) -> pd.DataFrame:
    """
    songs_df × events_df を結合して申告フォーマット DataFrame を返す。
    1行 per event（同じ曲が複数箇所で使われる場合はそれぞれ別行）。
    join key: イベント名
    """
    # songs_df の申告列 + 補助列（確認ステータス・委任者・放送・配信・CD名）を選択
    _extra_cols = [c for c in _SHINKOK_AUX_COLS if c in songs_df.columns]
    song_rename = {src: dst for src, dst in _SHINKOK_RENAME.items() if src in songs_df.columns}
    songs_sub = songs_df[["イベント名"] + list(song_rename.keys()) + _extra_cols].copy()
    songs_sub.rename(columns=song_rename, inplace=True)

    event_pick = ["イベント名", "START TIME", "使用尺", "使用時間（分）", "使用時間（秒）"]
    events_sub = events_df[[c for c in event_pick if c in events_df.columns]].copy()

    merged = events_sub.merge(songs_sub, on="イベント名", how="left")

    # 使用尺から使用時間（分・秒）を再計算（DBロード時に欠損していても補完）
    if "使用尺" in merged.columns:
        _times = merged["使用尺"].apply(duration_to_min_sec)
        merged["使用時間（分）"] = _times.apply(lambda x: x[0])
        merged["使用時間（秒）"] = _times.apply(lambda x: x[1])

    out_cols = [c for c in _SHINKOK_ORDER if c in merged.columns]
    result = merged[out_cols]

    # IN タイム昇順ソート（START TIME が "HH:MM:SS.xx" 形式の文字列でも辞書順で正しく並ぶ）
    if "START TIME" in result.columns:
        result = result.sort_values("START TIME", kind="stable").reset_index(drop=True)
    else:
        result = result.reset_index(drop=True)

    # 使用時間（分・秒）を整数型に
    for col in _SHINKOK_INT_COLS:
        if col in result.columns:
            result[col] = pd.to_numeric(result[col], errors="coerce").fillna(0).astype(int)

    # コード表示フォーマット
    if "JASRACコード" in result.columns:
        result["JASRACコード"] = result["JASRACコード"].apply(_fmt_jasrac_code)
    if "レコード番号" in result.columns:
        result["レコード番号"] = result["レコード番号"].apply(_fmt_cd_number)

    return result


def _write_shinkok_sheet(ws, df: pd.DataFrame) -> None:
    """申告フォーマットシートを書き込む。参照列はグレー背景で区別する。"""
    if df is None or len(df) == 0:
        ws.cell(row=1, column=1, value="データなし")
        return

    SHINKOK_HEADER_FILL = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
    REF_HEADER_FILL     = PatternFill(start_color="546E7A", end_color="546E7A", fill_type="solid")
    REF_CELL_FILL       = PatternFill(start_color="ECEFF1", end_color="ECEFF1", fill_type="solid")

    col_names = list(df.columns)

    for col_idx, col_name in enumerate(col_names, 1):
        fill = REF_HEADER_FILL if col_name in _SHINKOK_REF_COLS else SHINKOK_HEADER_FILL
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, (_, series) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(col_names, 1):
            val = series[col_name]
            if pd.isna(val):
                val = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DEFAULT_FONT
            if col_name in _SHINKOK_REF_COLS:
                cell.fill = REF_CELL_FILL

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    for col_idx, col_name in enumerate(col_names, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col_name))
        for r in range(2, min(ws.max_row, 200) + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v:
                max_len = max(max_len, min(len(str(v)), 40))
        ws.column_dimensions[col_letter].width = max_len + 2


def export_to_excel(
    songs_df: pd.DataFrame | None,
    events_df: pd.DataFrame | None,
    wav_df: pd.DataFrame | None,
    mp3_df: pd.DataFrame | None,
    search_df: pd.DataFrame | None,
    shinkok_df: pd.DataFrame | None = None,
) -> bytes:
    """
    各 DataFrame を Excel ファイルとして書き出す。
    Returns: Excel バイト列（st.download_button に直接渡せる）
    """
    wb = Workbook()
    wb.remove(wb.active)  # デフォルト空シートを削除

    if shinkok_df is not None and len(shinkok_df) > 0:
        _write_shinkok_sheet(wb.create_sheet("申告フォーマット"), shinkok_df)
    _write_sheet(wb.create_sheet("楽曲まとめ"), songs_df)
    _write_sheet(wb.create_sheet("イベント一覧"), events_df)
    _write_sheet(wb.create_sheet("WAV一覧"), wav_df)
    _write_sheet(wb.create_sheet("MP3一覧"), mp3_df)
    _write_sheet(wb.create_sheet("検索語一覧"), search_df)
    _write_notes_sheet(wb.create_sheet("確認メモ"))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
